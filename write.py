from openpyxl import Workbook

def ExportResult(bookList):
    wb = Workbook()
    ws = wb.active
    ws.title = "Chinese Books"

    ws['A1'] = 'No.'
    ws['B1'] = 'Title'
    ws['C1'] = 'Author'
    ws['D1'] = 'Isbn'
    ws['E1'] = 'Publisher'
    ws['F1'] = 'Edition'
    ws['G1'] = 'Page'
    ws['H1'] = 'Price'
    ws['I1'] = 'Package'
    for i in range(len(bookList)):
        ws['A%d' % (i+2)] = i+1
        ws['B%d' % (i+2)] = bookList[i].title
        ws['C%d' % (i+2)] = bookList[i].author
        ws['D%d' % (i+2)] = bookList[i].isbn
        ws['E%d' % (i+2)] = bookList[i].publisher
        ws['F%d' % (i+2)] = bookList[i].edition
        ws['G%d' % (i+2)] = bookList[i].page
        ws['H%d' % (i+2)] = bookList[i].price
        ws['I%d' % (i+2)] = bookList[i].package

    wb.save('document.xlsx')
