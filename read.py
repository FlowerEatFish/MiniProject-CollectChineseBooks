from openpyxl import load_workbook

class IsbnCollection():
    isbnList = [] # list

    def __init__(self):
        path = self.setPath()
        if path != "exit":
            fileData = self.getFullFile(path)
            self.isbnList = self.getIsbnList(fileData)
        else:
            print('Bye')

    def setPath(self):
        print('Please enter the path for .xlsx file:')
        print('That enter "exit" will end the command.')
        while True:
            inputPath = input()
            if self.isValidFile(inputPath):
                return inputPath
            if inputPath == "exit":
                return "exit"
            print('Error: invalid path.')

    def isValidFile(self, path):
        try:
            load_workbook(filename = path)
            return True
        except:
            return False

    def getFullFile(self, path):
        print('Collecting %s ...' % path)
        excelFile = load_workbook(filename = path)
        fileData = excelFile
        return fileData

    def getIsbnList(self, fileData):
        setColumn = self.getColumnForIsbn(fileData)
        sheet = fileData[fileData.get_sheet_names()[0]]
        getIsbnList = []
        noneData = 0
        for i in range(100):
            if noneData > 2:
                break
            getIsbn = sheet['%s%d' % (setColumn,(i+2))].value
            if getIsbn is None:
                noneData += 1
                getIsbnList.append('None')
            else:
                noneData = 0
                getIsbnList.append(getIsbn)
        return getIsbnList

    def getColumnForIsbn(self, fileData):
        sheet = fileData[fileData.get_sheet_names()[0]]
        for i in range(26):
            checkBlock = sheet['%s1' % chr(ord('A')+i)]
            if self.isIsbnFromColumn(checkBlock):
                return '%s' % chr(ord('A')+i)

    def isIsbnFromColumn(self, singleBlock):
        if singleBlock.value == 'ISBN':
            return True

# Demo
if __name__ == '__main__':
    result = IsbnCollection()
    print(result.isbnList)
