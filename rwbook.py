""" A library called openpyxl. Used for read/write excel file. """
from openpyxl import load_workbook, Workbook

class Read():
    """ Read an excel file and parse value of ISBN column. """
    isbn_list = [] # list

    def __init__(self):
        path = self.set_path()
        if path != "exit":
            file_data = self.get_full_file(path)
            self.isbn_list = self.get_isbn_list(file_data)
        else:
            print('Bye')
            input()

    def set_path(self):
        """ Submit file name or end program by enter 'exit'. Return string. """
        print('Please enter the path for .xlsx file:')
        print('That enter "exit" will end the command.')
        while True:
            input_path = input()
            if self.is_valid_file(input_path):
                return input_path
            if input_path == "exit":
                return "exit"

    @staticmethod
    def is_valid_file(path):
        """ Check whether file is present. Return boolean. """
        try:
            load_workbook(filename=path)
            return True
        except:
            print('Error: invalid path.')
            return False

    @staticmethod
    def get_full_file(path):
        """ Read content from file. """
        print('Collecting %s ...' % path)
        excel_file = load_workbook(filename=path)
        file_data = excel_file
        print('Done in collection.')
        return file_data

    def get_isbn_list(self, file_data):
        """ Parse value of ISBN column. """
        set_column = self.get_column_for_isbn(file_data)
        sheet = file_data[file_data.get_sheet_names()[0]]
        get_isbn_list = []
        none_data = 0
        for i in range(100):
            if none_data > 2:
                break
            get_isbn = sheet['%s%d' % (set_column, (i+2))].value
            if get_isbn is None:
                none_data += 1
                get_isbn_list.append('None')
            else:
                none_data = 0
                get_isbn_list.append(get_isbn)
        return get_isbn_list

    def get_column_for_isbn(self, file_data):
        """ Find ISBN column. Return string. """
        sheet = file_data[file_data.get_sheet_names()[0]]
        for i in range(26):
            check_block = sheet['%s1' % chr(ord('A')+i)]
            if self.is_isbn_from_column(check_block):
                return '%s' % chr(ord('A')+i)

    @staticmethod
    def is_isbn_from_column(single_block):
        """ Check whether specific column is ISBN. Return boolean. """
        if single_block.value == 'ISBN':
            return True
        return False

class Write():
    """ Write value collected from ISBN into an excel file. """
    def __init__(self, book_list):
        self.write_file(book_list)

    @staticmethod
    def write_file(book_list):
        """ Write value and create a file. """
        wb = Workbook()
        ws = wb.active
        ws.title = "Chinese Books"

        ws['A1'] = 'No.'
        ws['B1'] = 'Title'
        ws['C1'] = 'Author'
        ws['D1'] = 'Isbn'
        ws['E1'] = 'Publisher'
        ws['F1'] = 'Edition'
        ws['G1'] = 'Page'
        ws['H1'] = 'Price'
        ws['I1'] = 'Package'
        for i in range(len(book_list)):
            ws['A%d' % (i+2)] = i+1
            ws['B%d' % (i+2)] = book_list[i].title
            ws['C%d' % (i+2)] = book_list[i].author
            ws['D%d' % (i+2)] = book_list[i].isbn
            ws['E%d' % (i+2)] = book_list[i].publisher
            ws['F%d' % (i+2)] = book_list[i].edition
            ws['G%d' % (i+2)] = book_list[i].page
            ws['H%d' % (i+2)] = book_list[i].price
            ws['I%d' % (i+2)] = book_list[i].package

        wb.save('document.xlsx')

# Demo
if __name__ == '__main__':
    RESULT = Read()
    print(RESULT.isbn_list)
